from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, PatternFill

#configure input files and result file
#result file is a copy of INP_FILE2 with difference in cells filled with colour
INP_FILE1 = r"test_data/file1.xlsx"
INP_FILE2 = r"test_data/file2.xlsx"
RESULT_FILE = r"output/result_file.xlsx"

print(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

fill = PatternFill("solid", fgColor="00FFFF00") #configure the colour to fill in 

workbook1 = load_workbook(filename=INP_FILE1)
workbook2 = load_workbook(filename=INP_FILE2)
  
workbook1_sheet1 = workbook1.active
workbook2_sheet1 = workbook2.active

for i in range(1, max(workbook1_sheet1.max_row, workbook2_sheet1.max_row)+1):
    for j in range(1, max(workbook1_sheet1.max_column, workbook2_sheet1.max_column)+1):
        cell_obj1 = workbook1_sheet1.cell(row=i, column=j)
        value1 = cell_obj1.value
        cell_obj2 = workbook2_sheet1.cell(row=i, column=j)
        value2 = cell_obj2.value
        if value1 != value2:
            cell_obj2.fill = fill

workbook2.save(RESULT_FILE)

print(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
